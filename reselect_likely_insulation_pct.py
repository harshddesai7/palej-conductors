import math
import shutil
from datetime import datetime
from pathlib import Path
import itertools

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


MISSING_TOKENS = {"", "---", "--", "#VALUE!", "nan", "None"}

# Highest weightage to KG as requested.
WEIGHT_KG = 0.5
WEIGHT_SCRAP = 0.25
WEIGHT_MATCH = 0.25


def parse_num(value):
    if value is None:
        return None
    s = str(value).strip()
    if s in MISSING_TOKENS:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def normalize_minmax(vals):
    if not vals:
        return []
    vmin = min(vals)
    vmax = max(vals)
    if math.isclose(vmin, vmax):
        return [1.0 for _ in vals]
    return [(v - vmin) / (vmax - vmin) for v in vals]


def compute_match_scores(pcts):
    """
    Match score = closeness to group's central trend (median), normalized by range.
    Closer to peers => higher score.
    """
    if not pcts:
        return []
    if len(pcts) == 1:
        return [1.0]
    center = float(pd.Series(pcts).median())
    pmin = min(pcts)
    pmax = max(pcts)
    prange = pmax - pmin
    if math.isclose(prange, 0.0):
        return [1.0 for _ in pcts]

    # scale by half-range: center-ish values score high, outliers score low
    denom = prange / 2.0
    scores = []
    for p in pcts:
        dev = abs(p - center) / denom if denom > 0 else 0
        score = max(0.0, 1.0 - dev)
        scores.append(score)
    return scores


def select_row_for_group(group_df):
    """
    Return index of selected row for this size group.
    Selection score:
      0.5 * KG_score + 0.25 * low_scrap_score + 0.25 * match_score
    Tie-breakers:
      higher KG, lower scrap rate, closer to median, lower insulation %
    """
    rows = []
    for idx, row in group_df.iterrows():
        pct = parse_num(row.get("Insulation Per %"))
        if pct is None:
            continue

        kg = (
            parse_num(row.get("Actual Bare wt"))
            or parse_num(row.get("Final Dis.Qty."))
            or parse_num(row.get("Insulation wt kg"))
            or 0.0
        )
        scrap = parse_num(row.get("Scrap"))
        rows.append(
            {
                "idx": idx,
                "pct": pct,
                "kg": float(kg),
                "scrap": scrap,
            }
        )

    if not rows:
        return None, None
    if len(rows) == 1:
        return rows[0]["idx"], rows[0]["pct"]

    # Outlier guard: for 3+ duplicates, keep closest n-1 insulation % values.
    # This enforces "matches other values for same size" before scoring.
    if len(rows) >= 3:
        k = len(rows) - 1
        best_subset = None
        best_key = None
        for combo in itertools.combinations(rows, k):
            vals = [r["pct"] for r in combo]
            c_range = max(vals) - min(vals)
            c_mean = sum(vals) / len(vals)
            c_var = sum((v - c_mean) ** 2 for v in vals) / len(vals)
            c_min = min(vals)
            key = (c_range, c_var, c_min)
            if best_key is None or key < best_key:
                best_key = key
                best_subset = list(combo)
        rows = best_subset

    # KG score (higher is better)
    kg_vals = [r["kg"] for r in rows]
    kg_scores = normalize_minmax(kg_vals)

    # Scrap score using scrap rate (lower is better)
    raw_rates = []
    for r in rows:
        if r["scrap"] is None or r["kg"] <= 0:
            raw_rates.append(None)
        else:
            raw_rates.append(r["scrap"] / r["kg"])

    observed_rates = [x for x in raw_rates if x is not None]
    if observed_rates:
        worst = max(observed_rates)
        filled_rates = [x if x is not None else worst * 1.1 for x in raw_rates]
        rate_scores = normalize_minmax(filled_rates)
        # Lower rate should score higher
        scrap_scores = [1.0 - s for s in rate_scores]
    else:
        scrap_scores = [0.5 for _ in rows]
        filled_rates = [999999.0 for _ in rows]

    # Match score (closer to group trend is better)
    pct_vals = [r["pct"] for r in rows]
    match_scores = compute_match_scores(pct_vals)
    pct_median = float(pd.Series(pct_vals).median())

    # Final weighted score and selection
    best = None
    for i, r in enumerate(rows):
        final_score = (
            WEIGHT_KG * kg_scores[i]
            + WEIGHT_SCRAP * scrap_scores[i]
            + WEIGHT_MATCH * match_scores[i]
        )
        # tie-breakers (descending by score, kg), then ascending scrap rate, deviation, pct
        tiebreak = (
            final_score,
            r["kg"],
            -filled_rates[i],
            -abs(r["pct"] - pct_median),
            -r["pct"],
        )
        if best is None or tiebreak > best["tiebreak"]:
            best = {
                "idx": r["idx"],
                "pct": r["pct"],
                "tiebreak": tiebreak,
            }

    return best["idx"], best["pct"]


def apply_sheet_logic(df):
    out = df.copy()
    size_col = "Size Key" if "Size Key" in out.columns else "Size"

    # reset and refill recommendation columns
    if "Likely Insulation % Increase" not in out.columns:
        out["Likely Insulation % Increase"] = ""
    if "Recommended % Marked" not in out.columns:
        out["Recommended % Marked"] = ""

    out["Likely Insulation % Increase"] = ""
    out["Recommended % Marked"] = ""

    for size_key, grp in out.groupby(size_col, sort=False):
        selected_idx, selected_pct = select_row_for_group(grp)
        if selected_idx is None:
            continue
        selected_text = f"{selected_pct:.9f}".rstrip("0").rstrip(".")
        out.loc[grp.index, "Likely Insulation % Increase"] = selected_text
        out.loc[selected_idx, "Recommended % Marked"] = "Yes"

    return out


def apply_formatting(xlsx_path):
    wb = load_workbook(xlsx_path)
    yellow = PatternFill(fill_type="solid", start_color="FFFDE9D9", end_color="FFFDE9D9")
    green = PatternFill(fill_type="solid", start_color="FFC6EFCE", end_color="FFC6EFCE")
    no_fill = PatternFill(fill_type=None)

    for ws in wb.worksheets:
        headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        size_key_col = headers.get("Size Key")
        dup_col = headers.get("Duplicate?")
        mark_col = headers.get("Recommended % Marked")
        pct_col = headers.get("Insulation Per %")
        likely_col = headers.get("Likely Insulation % Increase")

        if not all([mark_col, pct_col, likely_col]):
            continue

        # clear previous fills on related columns
        for r in range(2, ws.max_row + 1):
            ws.cell(r, pct_col).fill = no_fill
            ws.cell(r, likely_col).fill = no_fill
            if size_key_col:
                ws.cell(r, size_key_col).fill = no_fill
            if dup_col:
                ws.cell(r, dup_col).fill = no_fill

        # apply fresh fills
        for r in range(2, ws.max_row + 1):
            mark_val = ws.cell(r, mark_col).value
            dup_val = ws.cell(r, dup_col).value if dup_col else None

            if dup_val == "Duplicate":
                if size_key_col:
                    ws.cell(r, size_key_col).fill = yellow
                if dup_col:
                    ws.cell(r, dup_col).fill = yellow

            if mark_val == "Yes":
                ws.cell(r, pct_col).fill = green
                ws.cell(r, likely_col).fill = green

        ws.freeze_panes = "A2"

    wb.save(xlsx_path)


def main():
    xlsx_path = Path(r"c:\Projects\Palej Calculation App\DFG_Data_Updated.xlsx")
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Workbook not found: {xlsx_path}")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = xlsx_path.with_suffix(f".xlsx.bak_{ts}")
    shutil.copy2(xlsx_path, backup_path)

    xls = pd.ExcelFile(xlsx_path)
    processed = {}
    for sheet in xls.sheet_names:
        df = pd.read_excel(xlsx_path, sheet_name=sheet, dtype=str, keep_default_na=False)
        processed[sheet] = apply_sheet_logic(df)

    temp_out = xlsx_path.with_name("DFG_Data_Updated_reweighted_tmp.xlsx")
    with pd.ExcelWriter(temp_out, engine="openpyxl") as writer:
        for sheet, df in processed.items():
            df.to_excel(writer, sheet_name=sheet, index=False)

    apply_formatting(temp_out)

    replaced = False
    try:
        shutil.move(temp_out, xlsx_path)
        replaced = True
    except Exception:
        replaced = False

    if replaced:
        print("Updated workbook:", xlsx_path)
    else:
        print("Workbook is locked, wrote updated file to:", temp_out)
    print("Backup created:", backup_path)
    for sheet, df in processed.items():
        marked = (df["Recommended % Marked"] == "Yes").sum()
        sizes = df["Size Key"].nunique() if "Size Key" in df.columns else df["Size"].nunique()
        print(f"{sheet}: selected_rows={marked}, unique_sizes={sizes}")


if __name__ == "__main__":
    main()
