import itertools
import math
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


MISSING_TOKENS = {"", "---", "--", "#VALUE!", "nan", "None"}

# Green selection criteria (same intent as previous):
# highest KG priority, then least scrap, then match to nearby values
GREEN_W_KG = 0.50
GREEN_W_SCRAP = 0.25
GREEN_W_MATCH = 0.25

# Top-5 factor reliability weighting
REL_W_KG = 0.30
REL_W_TOTAL = 0.25
REL_W_SCRAP = 0.20
REL_W_MATCH = 0.20
REL_W_COMPLETE = 0.05


def parse_num(value):
    if value is None:
        return None
    s = str(value).strip()
    if s in MISSING_TOKENS:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def normalize(vals):
    if not vals:
        return []
    mn = min(vals)
    mx = max(vals)
    if math.isclose(mn, mx):
        return [1.0 for _ in vals]
    return [(v - mn) / (mx - mn) for v in vals]


def build_size_key(df: pd.DataFrame) -> pd.Series:
    if {"Width", "Thickness"}.issubset(df.columns):
        return (
            df["Width"].astype(str).str.strip()
            + " x "
            + df["Thickness"].astype(str).str.strip()
        )
    if {"Wire Value", "Wire Unit"}.issubset(df.columns):
        return (
            df["Wire Value"].astype(str).str.strip()
            + " "
            + df["Wire Unit"].astype(str).str.upper().str.strip()
        )
    return df["Size"].astype(str).str.strip()


def pick_inlier_subset(rows):
    """
    rows: list of dicts with keys: idx, pct, kg, scrap
    For 3+ rows, use closest n-1 pct cluster to reduce one outlier.
    """
    n = len(rows)
    if n <= 2:
        return rows
    k = n - 1
    best_subset = None
    best_key = None
    for combo in itertools.combinations(rows, k):
        vals = [r["pct"] for r in combo]
        r_range = max(vals) - min(vals)
        mean = sum(vals) / len(vals)
        var = sum((v - mean) ** 2 for v in vals) / len(vals)
        key = (r_range, var, min(vals))
        if best_key is None or key < best_key:
            best_key = key
            best_subset = list(combo)
    return best_subset


def select_green_row(group_df: pd.DataFrame):
    """
    Return selected row index and likely pct for one Size Key group.
    """
    rows = []
    for idx, row in group_df.iterrows():
        pct = parse_num(row.get("Insulation Per %"))
        if pct is None:
            continue
        kg = (
            parse_num(row.get("Actual Bare wt"))
            or parse_num(row.get("Final Dis.Qty."))
            or parse_num(row.get("Insulation wt kg"))
            or 0.0
        )
        scrap = parse_num(row.get("Scrap"))
        rows.append({"idx": idx, "pct": pct, "kg": float(kg), "scrap": scrap})

    if not rows:
        return None, None
    if len(rows) == 1:
        return rows[0]["idx"], rows[0]["pct"]

    rows = pick_inlier_subset(rows)
    pcts = [r["pct"] for r in rows]
    center = float(pd.Series(pcts).median())
    p_range = max(pcts) - min(pcts)

    kg_scores = normalize([r["kg"] for r in rows])

    rates = []
    for r in rows:
        if r["scrap"] is None or r["kg"] <= 0:
            rates.append(None)
        else:
            rates.append(r["scrap"] / r["kg"])

    observed = [x for x in rates if x is not None]
    if observed:
        worst = max(observed)
        filled_rates = [x if x is not None else worst * 1.1 for x in rates]
        rate_norm = normalize(filled_rates)
        scrap_scores = [1.0 - x for x in rate_norm]  # lower scrap is better
    else:
        scrap_scores = [0.5 for _ in rows]

    match_scores = []
    for r in rows:
        if math.isclose(p_range, 0.0):
            match_scores.append(1.0)
        else:
            dev = abs(r["pct"] - center) / (p_range / 2.0)
            match_scores.append(max(0.0, 1.0 - dev))

    best = None
    for i, r in enumerate(rows):
        score = (
            GREEN_W_KG * kg_scores[i]
            + GREEN_W_SCRAP * scrap_scores[i]
            + GREEN_W_MATCH * match_scores[i]
        )
        # tie: higher kg, then lower pct
        key = (score, r["kg"], -r["pct"])
        if best is None or key > best["key"]:
            best = {"idx": r["idx"], "pct": r["pct"], "key": key}
    return best["idx"], best["pct"]


def process_sheet(df: pd.DataFrame):
    out = df.copy()
    out["Size Key"] = build_size_key(out)
    out["Duplicate Count"] = out.groupby("Size Key")["Size Key"].transform("count")
    out["Duplicate?"] = out["Duplicate Count"].apply(
        lambda n: "Duplicate" if int(n) > 1 else "Unique"
    )

    if "Likely Insulation % Increase" not in out.columns:
        out["Likely Insulation % Increase"] = ""
    if "Recommended % Marked" not in out.columns:
        out["Recommended % Marked"] = ""

    out["Likely Insulation % Increase"] = ""
    out["Recommended % Marked"] = ""

    for _, grp in out.groupby("Size Key", sort=False):
        selected_idx, likely_pct = select_green_row(grp)
        if selected_idx is None:
            continue
        likely_txt = f"{likely_pct:.9f}".rstrip("0").rstrip(".")
        out.loc[grp.index, "Likely Insulation % Increase"] = likely_txt
        out.loc[selected_idx, "Recommended % Marked"] = "Yes"

    return out


def build_reliability_rows(sheets):
    rows = []
    for sheet_name, df in sheets.items():
        for idx, row in df.iterrows():
            factor = parse_num(row.get("factor"))
            pct = parse_num(row.get("Insulation Per %"))
            likely = parse_num(row.get("Likely Insulation % Increase"))
            kg = parse_num(row.get("Actual Bare wt")) or 0.0
            total = parse_num(row.get("Final Dis.Qty.")) or 0.0
            scrap = parse_num(row.get("Scrap"))
            scrap_rate = (scrap / kg) if (scrap is not None and kg > 0) else None

            core_cols = [
                "Insulation Per %",
                "factor",
                "Actual Bare wt",
                "Final Dis.Qty.",
                "Scrap",
                "Likely Insulation % Increase",
            ]
            present = 0
            for c in core_cols:
                v = row.get(c, "")
                if str(v).strip() not in MISSING_TOKENS:
                    present += 1
            completeness = present / len(core_cols)

            rows.append(
                {
                    "sheet": sheet_name,
                    "idx": idx,
                    "size_key": str(row.get("Size Key", "")),
                    "factor": factor,
                    "pct": pct,
                    "likely": likely,
                    "kg": float(kg),
                    "total": float(total),
                    "scrap_rate": scrap_rate,
                    "completeness": completeness,
                }
            )
    return rows


def compute_top5_factor_labels(sheets):
    rows = build_reliability_rows(sheets)
    valid = [r for r in rows if r["factor"] is not None and r["pct"] is not None]
    if not valid:
        return {}, pd.DataFrame()

    kg_scores = normalize([r["kg"] for r in valid])
    total_scores = normalize([r["total"] for r in valid])

    rates = [r["scrap_rate"] for r in valid]
    observed = [x for x in rates if x is not None]
    if observed:
        worst = max(observed)
        filled = [x if x is not None else worst * 1.1 for x in rates]
        rate_norm = normalize(filled)
        scrap_scores = [1.0 - x for x in rate_norm]
    else:
        scrap_scores = [0.5 for _ in valid]

    # closeness to likely pct per row
    match_scores = []
    for r in valid:
        if r["likely"] is None:
            match_scores.append(0.5)
            continue
        base = max(1.0, abs(r["likely"]) * 0.25)
        dev = abs(r["pct"] - r["likely"]) / base
        match_scores.append(max(0.0, 1.0 - dev))

    bucket_step = 0.05
    for i, r in enumerate(valid):
        r["reliability"] = (
            REL_W_KG * kg_scores[i]
            + REL_W_TOTAL * total_scores[i]
            + REL_W_SCRAP * scrap_scores[i]
            + REL_W_MATCH * match_scores[i]
            + REL_W_COMPLETE * r["completeness"]
        )
        r["factor_bin"] = round(r["factor"] / bucket_step) * bucket_step

    # aggregate support by factor bin
    agg = {}
    for r in valid:
        k = r["factor_bin"]
        if k not in agg:
            agg[k] = {"support": 0.0, "count": 0, "kg_sum": 0.0, "scrap_rates": []}
        agg[k]["support"] += r["reliability"]
        agg[k]["count"] += 1
        agg[k]["kg_sum"] += r["kg"]
        if r["scrap_rate"] is not None:
            agg[k]["scrap_rates"].append(r["scrap_rate"])

    summary_rows = []
    for k, v in agg.items():
        avg_scrap = (
            sum(v["scrap_rates"]) / len(v["scrap_rates"]) if v["scrap_rates"] else None
        )
        summary_rows.append(
            {
                "factor_value": k,
                "support_score": round(v["support"], 6),
                "row_count": v["count"],
                "avg_kg": round(v["kg_sum"] / v["count"], 6) if v["count"] else 0,
                "avg_scrap_rate": round(avg_scrap, 8) if avg_scrap is not None else "",
            }
        )

    summary_df = pd.DataFrame(summary_rows).sort_values(
        ["support_score", "row_count"], ascending=[False, False]
    )
    top5 = summary_df.head(5)["factor_value"].tolist()
    rank_map = {f: i + 1 for i, f in enumerate(top5)}

    # map row -> label
    row_labels = {}
    row_reliability = {}
    for r in valid:
        key = (r["sheet"], r["idx"])
        row_reliability[key] = round(r["reliability"], 6)
        if r["factor_bin"] in rank_map:
            rank = rank_map[r["factor_bin"]]
            row_labels[key] = f"Top-{rank} ({r['factor_bin']:.2f})"
        else:
            row_labels[key] = ""

    # summary with rank
    summary_df["rank"] = summary_df["factor_value"].map(rank_map).fillna("")
    summary_df = summary_df.sort_values(
        by=["rank", "support_score"], ascending=[True, False], na_position="last"
    )
    return {"labels": row_labels, "reliability": row_reliability, "top5": top5}, summary_df


def apply_row_labels(sheets, label_data):
    labels = label_data["labels"]
    reliab = label_data["reliability"]
    for sheet_name, df in sheets.items():
        if "Top 5 Likely Factor" not in df.columns:
            df["Top 5 Likely Factor"] = ""
        if "Factor Reliability Score" not in df.columns:
            df["Factor Reliability Score"] = ""
        for idx in df.index:
            key = (sheet_name, idx)
            df.at[idx, "Top 5 Likely Factor"] = labels.get(key, "")
            val = reliab.get(key)
            df.at[idx, "Factor Reliability Score"] = "" if val is None else val
        sheets[sheet_name] = df
    return sheets


def apply_formatting(workbook_path: Path):
    wb = load_workbook(workbook_path)
    yellow = PatternFill(fill_type="solid", start_color="FFFDE9D9", end_color="FFFDE9D9")
    green = PatternFill(fill_type="solid", start_color="FFC6EFCE", end_color="FFC6EFCE")
    blue = PatternFill(fill_type="solid", start_color="FFD9E1F2", end_color="FFD9E1F2")
    no_fill = PatternFill(fill_type=None)

    for ws in wb.worksheets:
        if ws.title == "Factor_Top5_Summary":
            continue
        headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        size_key_col = headers.get("Size Key")
        dup_col = headers.get("Duplicate?")
        pct_col = headers.get("Insulation Per %")
        likely_col = headers.get("Likely Insulation % Increase")
        mark_col = headers.get("Recommended % Marked")
        factor_col = headers.get("factor")
        top5_col = headers.get("Top 5 Likely Factor")

        for r in range(2, ws.max_row + 1):
            # reset target fills first
            for c in [size_key_col, dup_col, pct_col, likely_col, factor_col, top5_col]:
                if c:
                    ws.cell(r, c).fill = no_fill

            dup_val = ws.cell(r, dup_col).value if dup_col else None
            mark_val = ws.cell(r, mark_col).value if mark_col else None
            top5_val = ws.cell(r, top5_col).value if top5_col else None

            if dup_val == "Duplicate":
                if size_key_col:
                    ws.cell(r, size_key_col).fill = yellow
                if dup_col:
                    ws.cell(r, dup_col).fill = yellow

            if mark_val == "Yes":
                if pct_col:
                    ws.cell(r, pct_col).fill = green
                if likely_col:
                    ws.cell(r, likely_col).fill = green

            if top5_val not in (None, ""):
                if factor_col:
                    ws.cell(r, factor_col).fill = blue
                if top5_col:
                    ws.cell(r, top5_col).fill = blue

        ws.freeze_panes = "A2"

    wb.save(workbook_path)


def main():
    path = Path(r"c:\Projects\Palej Calculation App\DFG_Data_Updated_reweighted_tmp.xlsx")
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = path.with_suffix(f".xlsx.bak_{ts}")
    shutil.copy2(path, backup_path)

    xls = pd.ExcelFile(path)
    source_sheets = [s for s in xls.sheet_names if s != "Factor_Top5_Summary"]
    sheets = {}
    for s in source_sheets:
        df = pd.read_excel(path, sheet_name=s, dtype=str).fillna("")
        sheets[s] = process_sheet(df)

    label_data, summary_df = compute_top5_factor_labels(sheets)
    sheets = apply_row_labels(sheets, label_data)

    out_path = path
    try:
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            for s in source_sheets:
                sheets[s].to_excel(writer, sheet_name=s, index=False)
            summary_df.to_excel(writer, sheet_name="Factor_Top5_Summary", index=False)
    except PermissionError:
        out_path = path.with_name("DFG_Data_Updated_reweighted_tmp_marked.xlsx")
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            for s in source_sheets:
                sheets[s].to_excel(writer, sheet_name=s, index=False)
            summary_df.to_excel(writer, sheet_name="Factor_Top5_Summary", index=False)

    apply_formatting(out_path)

    print("Workbook updated:", out_path)
    print("Backup created:", backup_path)
    for s in source_sheets:
        df = sheets[s]
        dup = (df["Duplicate?"] == "Duplicate").sum()
        green = (df["Recommended % Marked"] == "Yes").sum()
        top5 = df["Top 5 Likely Factor"].astype(str).str.strip().ne("").sum()
        print(f"{s}: rows={len(df)}, duplicate_rows={dup}, green_rows={green}, top5_factor_marked={top5}")
    print("Top 5 factor values:", label_data["top5"])


if __name__ == "__main__":
    main()
