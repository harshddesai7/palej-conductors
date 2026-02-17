import itertools
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


MISSING_TOKENS = {"", "---", "--", "#VALUE!", "nan", "None"}


def parse_pct(value) -> float | None:
    if value is None:
        return None
    s = str(value).strip()
    if s in MISSING_TOKENS:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def pick_recommended_cluster(values: list[tuple[int, float]]) -> tuple[float, set[int]]:
    """
    Pick most likely cluster and return:
      - recommended insulation % = lowest value in tightest cluster
      - row indices (within dataframe) that are part of that cluster

    Rule:
      - If 3+ values: choose subset of size n-1 with minimum range (drop one outlier),
        then take minimum of this closest subset.
      - If 1-2 values: closest subset is all values.
    """
    n = len(values)
    if n == 0:
        raise ValueError("values cannot be empty")

    if n <= 2:
        subset = values
    else:
        k = n - 1
        best_subset = None
        best_key = None
        for combo in itertools.combinations(values, k):
            nums = [v for _, v in combo]
            c_range = max(nums) - min(nums)
            c_min = min(nums)
            c_mean = sum(nums) / len(nums)
            c_var = sum((x - c_mean) ** 2 for x in nums) / len(nums)
            # Minimize range, then variance, then lower minimum
            key = (c_range, c_var, c_min)
            if best_key is None or key < best_key:
                best_key = key
                best_subset = list(combo)
        subset = best_subset

    rec_value = min(v for _, v in subset)
    subset_indices = {i for i, _ in subset}
    return rec_value, subset_indices


def build_size_key(df: pd.DataFrame) -> pd.Series:
    if {"Width", "Thickness"}.issubset(df.columns):
        return (
            df["Width"].astype(str).str.strip()
            + " x "
            + df["Thickness"].astype(str).str.strip()
        )
    if {"Wire Value", "Wire Unit"}.issubset(df.columns):
        return (
            df["Wire Value"].astype(str).str.strip()
            + " "
            + df["Wire Unit"].astype(str).str.upper().str.strip()
        )
    # Fallback
    return df["Size"].astype(str).str.strip()


def process_sheet(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["Size Key"] = build_size_key(out)
    out["Duplicate Count"] = out.groupby("Size Key")["Size Key"].transform("count")
    out["Duplicate?"] = out["Duplicate Count"].apply(
        lambda n: "Duplicate" if int(n) > 1 else "Unique"
    )

    pct_col = "Insulation Per %"
    out["_pct_num"] = out[pct_col].apply(parse_pct)
    out["Likely Insulation % Increase"] = ""
    out["Recommended % Marked"] = ""

    for size_key, group in out.groupby("Size Key", sort=False):
        valid = [(idx, val) for idx, val in group["_pct_num"].items() if val is not None]
        if not valid:
            continue

        rec_value, subset_indices = pick_recommended_cluster(valid)
        out.loc[group.index, "Likely Insulation % Increase"] = f"{rec_value:.9f}".rstrip(
            "0"
        ).rstrip(".")

        # Mark rows that match recommended value within tolerance (for that unique size)
        for idx, val in valid:
            if abs(val - rec_value) <= 1e-9:
                out.at[idx, "Recommended % Marked"] = "Yes"

    out = out.drop(columns=["_pct_num"])
    return out


def apply_formatting(path: Path) -> None:
    wb = load_workbook(path)
    yellow = PatternFill(fill_type="solid", start_color="FFFDE9D9", end_color="FFFDE9D9")
    green = PatternFill(fill_type="solid", start_color="FFC6EFCE", end_color="FFC6EFCE")

    for ws in wb.worksheets:
        headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        duplicate_col = headers.get("Duplicate?")
        size_key_col = headers.get("Size Key")
        pct_col = headers.get("Insulation Per %")
        likely_col = headers.get("Likely Insulation % Increase")
        marked_col = headers.get("Recommended % Marked")

        if not all([duplicate_col, size_key_col, pct_col, likely_col, marked_col]):
            continue

        for r in range(2, ws.max_row + 1):
            duplicate_val = ws.cell(r, duplicate_col).value
            marked_val = ws.cell(r, marked_col).value

            if duplicate_val == "Duplicate":
                ws.cell(r, size_key_col).fill = yellow
                ws.cell(r, duplicate_col).fill = yellow

            if marked_val == "Yes":
                ws.cell(r, pct_col).fill = green
                ws.cell(r, likely_col).fill = green

        ws.freeze_panes = "A2"

    wb.save(path)


def main() -> None:
    xlsx_path = Path(r"c:\Projects\Palej Calculation App\DFG_Data_Updated.xlsx")
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Workbook not found: {xlsx_path}")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = xlsx_path.with_suffix(f".xlsx.bak_{ts}")
    shutil.copy2(xlsx_path, backup_path)

    xls = pd.ExcelFile(xlsx_path)
    processed: dict[str, pd.DataFrame] = {}
    for sheet in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet, dtype=str, keep_default_na=False)
        processed[sheet] = process_sheet(df)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        for sheet, df in processed.items():
            df.to_excel(writer, sheet_name=sheet, index=False)

    apply_formatting(xlsx_path)

    print("Updated workbook:", xlsx_path)
    print("Backup created:", backup_path)
    for sheet, df in processed.items():
        dup_rows = (df["Duplicate?"] == "Duplicate").sum()
        marked_rows = (df["Recommended % Marked"] == "Yes").sum()
        print(f"{sheet}: rows={len(df)}, duplicate_rows={dup_rows}, green_marked_rows={marked_rows}")


if __name__ == "__main__":
    main()
