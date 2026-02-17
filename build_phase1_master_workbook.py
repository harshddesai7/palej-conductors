"""
Build one master workbook from phase-1 outputs:
- 4 tabs per insulation file (7 files => 28 tabs)
- keep only green-selected rows (Recommended % Marked == Yes)
- remove duplicates by unique size/insulation combination
- keep row factor values as-is
- highlight only top-3 likely factor bins in green
"""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

BASE = Path(r"c:\Projects\Palej Calculation App")
OUT_PATH = BASE / "Phase1_Master_Consolidated.xlsx"

SOURCE_FILES = [
    ("DFG", BASE / "DFG_Data.xlsx"),
    ("Poly", BASE / "Poly_Data_updated.xlsx"),
    ("PolyCotton", BASE / "PolyCotton_Data.xlsx"),
    ("PolyDFG", BASE / "PolyDFG_Data.xlsx"),
    ("PolyPaper", BASE / "PolyPaper_Data.xlsx"),
    ("EnamelDFG", BASE / "EnamelDFG_Data.xlsx"),
    ("Cotton", BASE / "Cotton_Data.xlsx"),
]

SOURCE_SHEETS = [
    "Aluminium Strips",
    "Copper Strips",
    "Aluminium Wires",
    "Copper Wires",
]

GREEN_FILL = PatternFill(fill_type="solid", start_color="FFC6EFCE", end_color="FFC6EFCE")
NO_FILL = PatternFill(fill_type=None)


def parse_num(value):
    if value is None:
        return None
    s = str(value).strip()
    if s in ("", "---", "--", "#VALUE!", "nan", "None"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def factor_bin(v: float, step: float = 0.05) -> float:
    return round(v / step) * step


def top3_factor_bins(summary_df: pd.DataFrame) -> list[float]:
    if summary_df is None or summary_df.empty:
        return []
    if "factor_value" not in summary_df.columns or "support_score" not in summary_df.columns:
        return []
    df = summary_df.copy()
    df["factor_value"] = pd.to_numeric(df["factor_value"], errors="coerce")
    df["support_score"] = pd.to_numeric(df["support_score"], errors="coerce")
    df = df.dropna(subset=["factor_value", "support_score"]).sort_values("support_score", ascending=False)
    return [float(x) for x in df.head(3)["factor_value"].tolist()]


def dedupe_green_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    if "Recommended % Marked" in df.columns:
        green = df[df["Recommended % Marked"].astype(str).str.strip() == "Yes"].copy()
    else:
        green = df.copy()

    if green.empty:
        return green

    # Unique combo key requested: size + insulation combination.
    size_key_col = "Size Key" if "Size Key" in green.columns else "Size"
    ins_col = "Type_of_Insulation" if "Type_of_Insulation" in green.columns else "Type of Insulation"
    if ins_col not in green.columns:
        ins_col = None

    if ins_col:
        green["_combo_key"] = green[size_key_col].astype(str).str.strip() + " | " + green[ins_col].astype(str).str.strip()
    else:
        green["_combo_key"] = green[size_key_col].astype(str).str.strip()

    # If duplicates remain, keep the strongest operational row:
    # highest Actual Bare wt, then lowest scrap rate.
    def row_score(r):
        kg = parse_num(r.get("Actual Bare wt")) or parse_num(r.get("Actual_Bare_Wt_kg")) or 0.0
        scrap = parse_num(r.get("Scrap"))
        rate = (scrap / kg) if (scrap is not None and kg > 0) else 1e9
        return (kg, -rate)

    green["_kg_score"] = green.apply(lambda r: row_score(r)[0], axis=1)
    green["_scrap_score"] = green.apply(lambda r: row_score(r)[1], axis=1)
    green = green.sort_values(["_combo_key", "_kg_score", "_scrap_score"], ascending=[True, False, False])
    green = green.drop_duplicates(subset=["_combo_key"], keep="first").copy()
    green = green.drop(columns=["_combo_key", "_kg_score", "_scrap_score"], errors="ignore")
    return green.reset_index(drop=True)


def sheet_out_name(insulation_name: str, base_sheet_name: str) -> str:
    # Keep short safe names for Excel sheet name length.
    suffix_map = {
        "Aluminium Strips": "Alu_Strips",
        "Copper Strips": "Cu_Strips",
        "Aluminium Wires": "Alu_Wires",
        "Copper Wires": "Cu_Wires",
    }
    return f"{insulation_name}_{suffix_map[base_sheet_name]}"


def mark_top3_factor_green(workbook_path: Path, sheet_name: str, top3_bins: Iterable[float]):
    bins = set(round(float(x), 6) for x in top3_bins)
    if not bins:
        return
    wb = load_workbook(workbook_path)
    ws = wb[sheet_name]
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    factor_col = headers.get("factor")
    if not factor_col:
        wb.save(workbook_path)
        return
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=factor_col)
        v = parse_num(cell.value)
        cell.fill = NO_FILL
        if v is None:
            continue
        b = round(factor_bin(v), 6)
        if b in bins:
            cell.fill = GREEN_FILL
    wb.save(workbook_path)


def main():
    sheets_to_write: list[tuple[str, pd.DataFrame, list[float]]] = []
    for insulation_name, path in SOURCE_FILES:
        if not path.exists():
            raise FileNotFoundError(f"Missing source workbook: {path}")
        xl = pd.ExcelFile(path)
        summary = pd.read_excel(path, sheet_name="Factor_Top5_Summary", dtype=str).fillna("") \
            if "Factor_Top5_Summary" in xl.sheet_names else pd.DataFrame()
        top3_bins = top3_factor_bins(summary)

        for src_sheet in SOURCE_SHEETS:
            df = pd.read_excel(path, sheet_name=src_sheet, dtype=str).fillna("")
            final_df = dedupe_green_rows(df)
            out_sheet = sheet_out_name(insulation_name, src_sheet)
            sheets_to_write.append((out_sheet, final_df, top3_bins))

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        for out_sheet, df, _ in sheets_to_write:
            df.to_excel(writer, sheet_name=out_sheet, index=False)

    for out_sheet, _, top3 in sheets_to_write:
        mark_top3_factor_green(OUT_PATH, out_sheet, top3)

    print(f"Created: {OUT_PATH}")
    print(f"Total tabs: {len(sheets_to_write)}")
    for out_sheet, df, top3 in sheets_to_write:
        print(f"{out_sheet}: rows={len(df)}, top3_factor_bins={top3}")


if __name__ == "__main__":
    main()
